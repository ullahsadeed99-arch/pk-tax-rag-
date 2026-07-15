"""Exports data/processed/finetune_raw.jsonl to an .xlsx workbook for manual
review. This is a review artifact only -- the actual OpenAI fine-tuning
upload uses data/processed/finetune_dataset.jsonl (see generate_finetune_data.py),
since OpenAI's fine-tuning API requires JSONL, not Excel.

Usage: python scripts/export_finetune_excel.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.config import PROCESSED_DIR

RAW_PATH = PROCESSED_DIR / "finetune_raw.jsonl"
XLSX_PATH = PROCESSED_DIR / "finetune_dataset.xlsx"

COLUMNS = ["doc_name", "section", "question", "answer", "chunk_id"]


def main() -> None:
    if not RAW_PATH.exists():
        raise SystemExit(f"{RAW_PATH} not found. Run scripts/generate_finetune_data.py first.")

    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs"
    ws.append(COLUMNS)

    n = 0
    with open(RAW_PATH, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row = json.loads(line)
            ws.append([row.get(c, "") for c in COLUMNS])
            n += 1

    widths = {"doc_name": 30, "section": 10, "question": 60, "answer": 80, "chunk_id": 28}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[col]
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"Wrote {n} rows to {XLSX_PATH}")


if __name__ == "__main__":
    main()
